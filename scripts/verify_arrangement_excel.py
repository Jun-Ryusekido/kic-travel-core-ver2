#!/usr/bin/env python3
# 手配書Excel出力の構造検証。
# - 結合セル範囲に重複がないか
# - 列幅がテンプレートと一致しているか
# - 行の高さが期待するブロックパターン(ヘッダー/通常日/最終日/フッター)通りか
# を、生成された各出力ファイルについてチェックする。
import sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

TEMPLATE = 'public/tour_arrangement_template.xlsx'
OUTPUTS = {
    3: 'scripts/tmp_template/out_3days.xlsx',
    11: 'scripts/tmp_template/out_11days.xlsx',
    14: 'scripts/tmp_template/out_14days.xlsx',
}

def ranges_overlap(a, b):
    a1,a2,a3,a4 = a
    b1,b2,b3,b4 = b
    return not (a3 < b1 or b3 < a1 or a4 < b2 or b4 < a2)

def check_merge_overlaps(ws, label):
    boxes = [range_boundaries(str(m)) for m in ws.merged_cells.ranges]
    problems = []
    for i in range(len(boxes)):
        for j in range(i+1, len(boxes)):
            if ranges_overlap(boxes[i], boxes[j]):
                problems.append((str(list(ws.merged_cells.ranges)[i]), str(list(ws.merged_cells.ranges)[j])))
    if problems:
        print(f'  [NG] {label}: 重複する結合範囲 {len(problems)}件')
        for p in problems[:10]:
            print('     ', p)
        return False
    print(f'  [OK] {label}: 結合範囲の重複なし ({len(boxes)}件)')
    return True

def check_col_widths(ws_t, ws_o, label):
    ok = True
    for col, dim_t in ws_t.column_dimensions.items():
        dim_o = ws_o.column_dimensions.get(col)
        wt = dim_t.width
        wo = dim_o.width if dim_o else None
        if wt is not None and (wo is None or abs(wt - wo) > 0.01):
            print(f'  [NG] {label}: 列{col}幅不一致 template={wt} output={wo}')
            ok = False
    if ok:
        print(f'  [OK] {label}: 列幅はテンプレートと一致')
    return ok

def check_row_heights(ws, n_days, label):
    ok = True
    expect = {}
    expect[1] = 18
    for r in range(2,10): expect[r] = 8.15
    normal_count = n_days - 1
    row = 11
    for _ in range(normal_count):
        for i in range(6): expect[row+i] = 11.15
        row += 6
    for i in range(6): expect[row+i] = 11.15
    row += 6
    for i in range(9): expect[row+i] = 15
    row += 9
    for r, h in expect.items():
        actual = ws.row_dimensions[r].height
        if actual is None or abs(actual - h) > 0.05:
            print(f'  [NG] {label}: 行{r}高さ不一致 expect={h} actual={actual}')
            ok = False
    if ok:
        print(f'  [OK] {label}: 行の高さがブロックパターン通り (最終フッター行={row-1})')
    return ok

def dump_sample_cells(ws, label):
    print(f'  --- {label} サンプルセル ---')
    for addr in ['A2','G5','O2','S2','S6','Z2','Z7','A11','B11','C11','E11','O11','R11','X11','Z11','Z13','Z15']:
        print(f'    {addr} = {ws[addr].value!r}')

def main():
    wb_t = load_workbook(TEMPLATE)
    ws_t = wb_t['F']
    all_ok = True
    for n, path in OUTPUTS.items():
        print(f'=== {n}日ツアー ({path}) ===')
        wb_o = load_workbook(path)
        ws_o = wb_o['F'] if 'F' in wb_o.sheetnames else wb_o.worksheets[0]
        ok1 = check_merge_overlaps(ws_o, f'{n}日')
        ok2 = check_col_widths(ws_t, ws_o, f'{n}日')
        ok3 = check_row_heights(ws_o, n, f'{n}日')
        dump_sample_cells(ws_o, f'{n}日')
        all_ok = all_ok and ok1 and ok2 and ok3
        print()
    sys.exit(0 if all_ok else 1)

if __name__ == '__main__':
    main()
